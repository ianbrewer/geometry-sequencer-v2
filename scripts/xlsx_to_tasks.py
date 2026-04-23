"""Import edits made in Task list.xlsx back into tasks.tsv.

Usage: python3 scripts/xlsx_to_tasks.py
Run this after editing the xlsx in Excel/Numbers so the tsv (canonical source)
picks up your changes. Review `git diff tasks.tsv` before committing.
"""
from __future__ import annotations

import sys
from datetime import datetime

import openpyxl

from tasks_schema import COLUMNS, TSV_PATH, XLSX_PATH, sort_rows, write_tsv

XLSX_HEADERS = [
    "Task ID",
    "Name",
    "Status",
    "Priority",
    "Issue Type",
    "App Section",
    "Owner",
    "Parent Task",
    "Notes",
    "Created At",
    "Completed At",
]
HEADER_TO_COL = {
    "Task ID": "task_id",
    "Name": "name",
    "Status": "status",
    "Priority": "priority",
    "Issue Type": "issue_type",
    "App Section": "app_section",
    "Owner": "owner",
    "Parent Task": "parent",
    "Notes": "notes",
    "Created At": "created_at",
    "Completed At": "completed_at",
}


def cell_to_str(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.isoformat(sep=" ", timespec="seconds")
    return str(v).strip()


def main():
    if not XLSX_PATH.exists():
        sys.exit(f"{XLSX_PATH} not found. Generate it first with tasks_to_xlsx.py.")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Tasks"]

    headers = [cell_to_str(c.value) for c in ws[1]]
    missing = [h for h in XLSX_HEADERS if h not in headers]
    if missing:
        sys.exit(f"Missing headers in xlsx: {missing}")
    idx = {h: headers.index(h) for h in XLSX_HEADERS}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        name = cell_to_str(row[idx["Name"]])
        if not name:
            continue
        rec = {col: "" for col in COLUMNS}
        for header, col in HEADER_TO_COL.items():
            rec[col] = cell_to_str(row[idx[header]])
        rows.append(rec)

    rows = sort_rows(rows)
    write_tsv(rows)
    print(f"Wrote {TSV_PATH} ({len(rows)} tasks)")


if __name__ == "__main__":
    main()
