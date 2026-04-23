"""Generate Task list.xlsx from tasks.tsv with dropdown validations.

Usage: python3 scripts/tasks_to_xlsx.py
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from tasks_schema import (
    APP_SECTION_OPTIONS,
    DROPDOWNS,
    ISSUE_TYPE_OPTIONS,
    OWNER_OPTIONS,
    PRIORITY_OPTIONS,
    STATUS_OPTIONS,
    XLSX_PATH,
    read_tsv,
    sort_rows,
)

XLSX_HEADERS = [
    "Task ID",
    "Name",
    "Status",
    "Priority",
    "Issue Type",
    "App Section",
    "Owner",
    "Parent Task",
    "Notes",
    "Created At",
    "Completed At",
]
XLSX_TO_TSV_KEY = {
    "Task ID": "task_id",
    "Name": "name",
    "Status": "status",
    "Priority": "priority",
    "Issue Type": "issue_type",
    "App Section": "app_section",
    "Owner": "owner",
    "Parent Task": "parent",
    "Notes": "notes",
    "Created At": "created_at",
    "Completed At": "completed_at",
}
DROPDOWN_COL_LETTERS = {
    "status": "C",
    "priority": "D",
    "issue_type": "E",
    "app_section": "F",
    "owner": "G",
}


def build_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"

    ws.append(XLSX_HEADERS)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F3E4D", end_color="2F3E4D", fill_type="solid")
    thin = Side(border_style="thin", color="D0D5DB")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, len(XLSX_HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = header_border
    ws.row_dimensions[1].height = 24

    for r in rows:
        ws.append([r.get(XLSX_TO_TSV_KEY[h], "") for h in XLSX_HEADERS])

    widths = {"A": 18, "B": 55, "C": 16, "D": 12, "E": 16, "F": 20, "G": 16, "H": 40, "I": 60, "J": 20, "K": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    last_row = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=len(XLSX_HEADERS)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"

    validation_end_row = max(last_row, 500)

    def add_dv(options, col_letter, label):
        formula = '"' + ",".join(options) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        dv.error = f"Pick one of: {', '.join(options)}"
        dv.errorTitle = f"Invalid {label}"
        dv.prompt = f"Select {label}"
        dv.promptTitle = label
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{validation_end_row}")

    add_dv(STATUS_OPTIONS, "C", "Status")
    add_dv(PRIORITY_OPTIONS, "D", "Priority")
    add_dv(ISSUE_TYPE_OPTIONS, "E", "Issue Type")
    add_dv(APP_SECTION_OPTIONS, "F", "App Section")
    add_dv(OWNER_OPTIONS, "G", "Owner")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(XLSX_HEADERS))}{last_row}"

    legend = wb.create_sheet("Legend")
    legend.append(["Field", "Allowed values"])
    for cell in (legend["A1"], legend["B1"]):
        cell.font = header_font
        cell.fill = header_fill
    for field, options in DROPDOWNS.items():
        legend.append([field, ", ".join(options)])
    legend.column_dimensions["A"].width = 16
    legend.column_dimensions["B"].width = 90
    for row in legend.iter_rows(min_row=2, max_row=legend.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    readme = wb.create_sheet("README")
    readme["A1"] = "This file is generated from tasks.tsv. Edits here can be round-tripped back via scripts/xlsx_to_tasks.py."
    readme["A2"] = "Canonical edits happen in tasks.tsv. Regenerate this workbook with: python3 scripts/tasks_to_xlsx.py"
    readme.column_dimensions["A"].width = 120
    for r in (1, 2):
        readme[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")

    return wb


def main():
    rows = sort_rows(read_tsv())
    wb = build_workbook(rows)
    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH} ({len(rows)} tasks)")


if __name__ == "__main__":
    main()
